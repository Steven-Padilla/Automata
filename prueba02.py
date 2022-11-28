import string
import os
from openpyxl import Workbook
from datetime import datetime
import PySimpleGUI as sg
from PySimpleGUI.PySimpleGUI import FolderBrowse, Listbox

#Variables for automata
finalState="z"
a=list(string.ascii_uppercase)
states=list(string.ascii_lowercase)
alphabet=[ 
   ['0'],['1'],['2'],['3'],['4'],['5'],['6'],['7'],['8'],['9'],[' '],['.'],['_'],['p'],['d'],['f'],['c'],['a'],['A'],['B'],['C'],['D'],['E'],['F'],['G'],['H'],['I'],['J'],['K'],['L'],['M'],['N'],["Ñ"],['O'],['P'],['Q'],['R'],['S'],['T'],['U'],['V'],['W'],['X'],['Y'],['Z']] 
transitions=[
#    0     1     2     3     4     5     6     7     8     9    10     11    12    13    14    15    16    17   18    19    20    21    22    23    24    25    26    27    28    29    30    31    32    33    34    35    36    37    38    39    40    41    42    43    44
 [ []   ,["b"],["c"],[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ] #0 - a
,[ []   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,["d"],["d"],[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ] #1 - b
,[ ["d"],["d"],["d"],[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ] #2 - c
,[ []   ,["e"],["e"],["e"],[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ] #3 - d
,[ ["f"],["f"],["f"],["f"],["f"],["f"],["f"],["f"],["f"],["f"],[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ] #3 - e
,[ ["g"],["g"],["g"],["g"],["g"],["g"],["g"],["g"],["g"],["g"],[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ] #5 - f
,[ ["h"],["h"],["h"],["h"],["h"],["h"],["h"],["h"],["h"],["h"],[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ] #6 - g
,[ []   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,["i"],[]   ,["i"],[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ] #7 - h
,[ []   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,["j"],["j"],["j"],["j"],["j"],["j"],["j"],["j"],["j"],["j"],["j"],["j"],["j"],["j"],["j"],["j"],["j"],["j"],["j"],["j"],["j"],["j"],["j"],["j"],["j"],["j"],["j"]] #8 - i
,[ []   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,["k"],["k"],["k"],["k"],["k"],["k"],["k"],["k"],["k"],["k"],["k"],["k"],["k"],["k"],["k"],["k"],["k"],["k"],["k"],["k"],["k"],["k"],["k"],["k"],["k"],["k"],["k"]] #9 - j
,[ []   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,["l"],["l"],["l"],["l"],["l"],["l"],["l"],["l"],["l"],["l"],["l"],["l"],["l"],["l"],["l"],["l"],["l"],["l"],["l"],["l"],["l"],["l"],["l"],["l"],["l"],["l"],["l"]] #10 - k
,[ []   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,["m"],[]   ,["m"],[]   ,[]   ,[]   ,[]   ,[]   ,["l"],["l"],["l"],["l"],["l"],["l"],["l"],["l"],["l"],["l"],["l"],["l"],["l"],["l"],["l"],["l"],["l"],["l"],["l"],["l"],["l"],["l"],["l"],["l"],["l"],["l"],["l"]] #11 - l
,[ []   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,["n"],["n"],["n"],["n"],["n"],["n"],["n"],["n"],["n"],["n"],["n"],["n"],["n"],["n"],["n"],["n"],["n"],["n"],["n"],["n"],["n"],["n"],["n"],["n"],["n"],["n"],["n"]] #12 - m
,[ []   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,["o"],["o"],["o"],["o"],["o"],["o"],["o"],["o"],["o"],["o"],["o"],["o"],["o"],["o"],["o"],["o"],["o"],["o"],["o"],["o"],["o"],["o"],["o"],["o"],["o"],["o"],["o"]] #13 - n
,[ []   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,["p"],["p"],["p"],["p"],["p"],["p"],["p"],["p"],["p"],["p"],["p"],["p"],["p"],["p"],["p"],["p"],["p"],["p"],["p"],["p"],["p"],["p"],["p"],["p"],["p"],["p"],["p"]] #14 - o
,[ []   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,["q"],["q"],["q"],[]   ,[]   ,[]   ,[]   ,[]   ,["p"],["p"],["p"],["p"],["p"],["p"],["p"],["p"],["p"],["p"],["p"],["p"],["p"],["p"],["p"],["p"],["p"],["p"],["p"],["p"],["p"],["p"],["p"],["p"],["p"],["p"],["p"]] #15 - p
,[ []   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,["r"],[]   ,[]   ,[]   ,["r"],[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ] #16 - q
,[ []   ,["s"],["s"],["s"],[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ] #17 - r
,[ []   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,["t"],["t"],["t"],[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ] #18 - s
,[ []   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,["u"],["u"],[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ] #29 - t
,[ []   ,["v"],["v"],["v"],["v"],["v"],["v"],["v"],["v"],["v"],[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ] #20 - u
,[ []   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,["w"],[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ] #21 - v
,[ []   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,["x"],[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ] #22 - w
,[ []   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,["y"],[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ] #23 - x
,[ []   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,["z"],[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ] #24 - y
,[ []   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ,[]   ]]#25 - z


#Variables for results
acceptedFiles=[]
notAcceptedFiles=[]

def errorM():
    layout =[
        [
            sg.Text('Error, folder vacio'),
        ]
    ]
    window = sg.Window("Error", layout)
    while True:
        event, values = window.read()
        if event == "Exit" or event == sg.WIN_CLOSED:
            break
def successM():
    layout =[
        [
            sg.Text('Ok, Report created'),
        ]
    ]
    window = sg.Window("Success", layout)
    while True:
        event, values = window.read()
        if event == "Exit" or event == sg.WIN_CLOSED:
            break
def window():
    file_list_column = [
        [
            sg.In(size=(50, 1), enable_events=True, key="-FOLDER-"),
            sg.FolderBrowse(),
            sg.Ok(key="CheckFolder")
        ],
        [
            sg.Listbox(
                values=[], enable_events=False, size=(50, 30),
                key="-FILE LIST-"
            )
        ]
    ]
    # image_viewer_column = [
        # [sg.Text("Elija la imagen que quiere visualizar:")],
        # [sg.Text(size=(30,1), key="-TOUT-")],
        # [sg.Image(key="-IMAGE-")],
        # []
    # ]
    layout = [
        [
            sg.Column(file_list_column),
            sg.VSeparator(),
            # sg.Column(image_viewer_column)
        ]
    ]
    window = sg.Window("Visualizador", layout)
    while True:
        event, values = window.read()
        if event == "Exit" or event == sg.WIN_CLOSED:
            break
        if event == "-FOLDER-":
            folder = values["-FOLDER-"]
            try:
                file_list = os.listdir(folder)
            except:
                file_list =[]

            fnames = [
                f
                for f in file_list
                if os.path.isfile(os.path.join(folder, f))
                # and f.lower().endswith((".png",".gif"))
            ]


            window["-FILE LIST-"].update(fnames)
        elif event == "-FILE LIST-":
            try:
                filename = os.path.join(
                    values["-FOLDER-"], values["-FILE LIST-"][0]
                )
                window["-TOUT-"].update(filename)
                window["-IMAGE-"].update(filename=filename)
            except:
                pass
        elif event =="CheckFolder":
            try:
                file_list = os.listdir(folder)
            except:
                file_list =[]
            fnames = [
                f
                for f in file_list
                if os.path.isfile(os.path.join(folder, f))
                # and f.lower().endswith((".png",".gif"))
            ]
            print("Checando folder")
            return fnames
            # getContent(fnames)

            break
    window.close()


def validateChar(simbol):
    index=0
    for aux in alphabet:
        for data in aux:
            if data == simbol:
                return index
        index+=1
    return -1


def validateState(state):
    index=0
    for singleState in states:
        if singleState==state:
            return index
        index+=1
    return -1

#---------Variables and constans for split dir name -------------
statesMatricula=["a","b","c","d","e","f","g"]
acceptedMatriculas=[]

statesLastname=["i","j","k","l"]
acceptedLastnames=[]

statesSecondLN=["m","n","o","p"]
acceptedSecondLN=[]

stateCorte="r"
cortes=[]

stateAct="u"
acts=[]
#----------------These are use down here ------------------------
def separator(state,char):
    if state in statesMatricula:
        return 0
    if state in statesLastname and (char !=" " and char!="_" and char != "."):        
        return 1
    if state in statesSecondLN and (char !=" " and char!="_" and char != "."):
        return 2
    if state == stateCorte:
        return 3
    if state == stateAct:
        return 4

    return -1


def validateWord(sentence):

    state="a"
    accepted=[]
    matricula=""
    lastname=""
    secondLN=""
    corte=""
    act=""
    for char in sentence:
        x=validateState(state)
        y=validateChar(char)

        
        if x == -1 or y==-1:
            notAcceptedFiles.append("".join(sentence))
            return accepted
        else:
            band=separator(state,char)
            if  band==0:
                matricula+=char
            elif band == 1:
                lastname+=char
            elif band==2:
                secondLN+=char
            elif band==3:
                corte+=char
            elif band==4:
                act+=char


            if len(transitions[x][y]) == 1:
                state=transitions[x][y][0]
            else:
                notAcceptedFiles.append("".join(sentence))
                return accepted
    if state==finalState:
        acceptedFiles.append("".join(sentence))
        acceptedMatriculas.append(matricula)
        acceptedLastnames.append(lastname)
        acceptedSecondLN.append(secondLN)
        cortes.append(corte)
        acts.append(act)
        return accepted

    else:
        notAcceptedFiles.append("".join(sentence))
        return accepted

def getContent(dir1):
    try:
        dirString=('./'+dir1)
        content=os.listdir(dirString)
        return content
    except:
        print('Folder not found')
        return 0


def generateReport():
    initialColumns=["Matriculas","Apellido1","Apellido2","Corte","Actividad","","","NO ACEPTADOS"]
    wb= Workbook()
    ws1=wb.active
    ws1.title="Datos"
    aux=2
    ws1.append(initialColumns)

    for matr in acceptedMatriculas:
        index=aux-2
        _=ws1.cell(column=1,row=aux,value=matr)
        _=ws1.cell(column=2,row=aux,value=acceptedLastnames[index])
        _=ws1.cell(column=3,row=aux,value=acceptedSecondLN[index])
        _=ws1.cell(column=4,row=aux,value=cortes[index])
        _=ws1.cell(column=5,row=aux,value=acts[index])
        aux+=1

    aux=2
    for single in notAcceptedFiles:
        _=ws1.cell(column=8,row=aux,value=single)
        aux+=1
    now = datetime.now() 
    format = now.strftime('%d-%m-%Y-%H-%M-%S')
    wb.save(filename=f'Report-{str(format)}.xlsx')

def main():
    # print('Drop the folder on the current directory')
    # dir1= input('Write your folder name: ')
    # content=getContent(dir1)
    content=window()
    if len(content)==0:
        return errorM()
    if content.__class__!=int:

        for name in content:
            word=validateWord(name)
        #Comment this if you only want the style sheet
        # print(f'Files accepted: \n\t{acceptedFiles}')
        # print(f'Files NOT accepted: \n\t{notAcceptedFiles}')
        #--------------------------------------------------#

        generateReport()
        successM()


main()
